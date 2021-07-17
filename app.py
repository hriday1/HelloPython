import openpyxl as xl
from openpyxl.chart import Reference, BarChart

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
print(cell.value)
for row in range(2, sheet.max_row+1 ):
    ch = chr(row+96)
    cell = sheet[f"c{row}"]
    print(f"c{row}")
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet[f'd{row}']
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')